"""
Scrapes all three AgencyZoom reports in one run and writes them to a single
Excel file with one sheet per report:

  Sales Report   — P&C Sales by Producer
  Monthly Report — Monthly Sales Summary
  Annual Report  — Annual Sales Summary

All shared logic (browser detection, login gate, scrape functions, Excel
helpers) lives in scraper_core.

Output: agencyzoom_reports.xlsx
"""

from pathlib import Path

from openpyxl import Workbook

from scraper_core import (
    scraper_session,
    scrape_sales, scrape_monthly, scrape_annual,
    write_sheet,
)

SCRAPER_PROFILE = Path(__file__).parent / "scraper_profile"
OUTPUT_FILE     = Path(__file__).parent / "agencyzoom_reports.xlsx"


def main():
    with scraper_session(SCRAPER_PROFILE) as (page, _):
        print("\nScraping all reports...")
        sales_h,   sales_rows                  = scrape_sales(page)
        print(f"  Sales: {len(sales_rows)} producer(s).")
        monthly_h, monthly_rows                = scrape_monthly(page)
        print(f"  Monthly: {len(monthly_rows)} month(s).")
        annual_h,  annual_rows, annual_total_i = scrape_annual(page)
        print(f"  Annual: {len(annual_rows)} staff row(s).")

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb.create_sheet("Sales Report"),   sales_h,   sales_rows)
    write_sheet(wb.create_sheet("Monthly Report"), monthly_h, monthly_rows)
    write_sheet(wb.create_sheet("Annual Report"),  annual_h,  annual_rows, annual_total_i)
    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
