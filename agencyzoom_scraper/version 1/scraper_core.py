"""
Shared core for the AgencyZoom scrapers — browser detection, a login-aware
session context manager, the per-page scrape functions, and Excel writing
helpers.

Every entry-point script imports from here:

  scrape_all_reports.py      — one xlsx, one sheet per report
  scrape_lead_pipelines.py   — one xlsx, one sheet per lead pipeline
  scrape_service_pipelines.py — one xlsx, one sheet per service pipeline
  lil_scraper.py             — GUI; one xlsx with three combined sheets

Each entry-point owns only its output filename and how it lays out the
workbook. All shared logic lives here.
"""

import re
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse, parse_qs

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


# ── browser detection ────────────────────────────────────────────────────────

BROWSER_PATHS = [
    ("Brave",  r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"),
    ("Brave",  r"C:\Program Files (x86)\BraveSoftware\Brave-Browser\Application\brave.exe"),
    ("Chrome", r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
    ("Chrome", r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
    ("Edge",   r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"),
    ("Edge",   r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
]


def detect_browser():
    """Return (name, executable_path) for the first installed Brave/Chrome/Edge."""
    for name, path in BROWSER_PATHS:
        if Path(path).exists():
            return name, path
    raise RuntimeError(
        "No supported browser found. Please install Brave, Chrome, or Edge."
    )


# ── session management ──────────────────────────────────────────────────────

AGENCYZOOM_HOME    = "https://app.agencyzoom.com/"
AGENCYZOOM_LOGOUT  = "https://app.agencyzoom.com/site/logout"
LOGIN_URL_KEYWORDS = ("login", "auth", "signin", "sign-in")


def default_login_prompt(browser_name):
    """CLI login gate. GUI front-ends pass their own dialog-based prompt."""
    print("\n" + "=" * 60)
    print(f" Log in to AgencyZoom in the {browser_name} window that just opened.")
    print(" Once you see the dashboard, return here and press Enter.")
    print("=" * 60)
    input(" Press Enter once logged in... ")


@contextmanager
def scraper_session(profile_dir, login_prompt=default_login_prompt, status_cb=print):
    """Launch a persistent browser, ensure login, yield (page, browser_name).

    On exit: logs out of AgencyZoom and closes every tab cleanly before
    closing the context, so Chromium has no session to restore next time.
    """
    profile_dir = Path(profile_dir)
    profile_dir.mkdir(exist_ok=True)
    browser_name, browser_exe = detect_browser()

    with sync_playwright() as p:
        status_cb(f"Launching {browser_name} (scraper profile: {profile_dir})...")
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(profile_dir),
            executable_path=browser_exe,
            headless=False,
            args=["--no-first-run", "--no-default-browser-check"],
        )

        # Trim any tabs Chromium tried to restore from a previous session.
        pages = list(context.pages)
        if pages:
            page = pages[0]
            for extra in pages[1:]:
                try:
                    extra.close()
                except Exception:
                    pass
        else:
            page = context.new_page()

        # Login gate
        page.goto(AGENCYZOOM_HOME, wait_until="domcontentloaded", timeout=60_000)
        page.wait_for_timeout(1500)
        if any(k in page.url.lower() for k in LOGIN_URL_KEYWORDS):
            login_prompt(browser_name)

        try:
            yield page, browser_name
        finally:
            status_cb("Logging out...")
            try:
                page.goto(AGENCYZOOM_LOGOUT,
                          wait_until="domcontentloaded", timeout=30_000)
            except Exception as e:
                status_cb(f"WARNING: Logout failed: {e}")
            for pg in list(context.pages):
                try:
                    pg.close()
                except Exception:
                    pass
            try:
                context.close()
            except Exception:
                pass


# ── report scrapers ─────────────────────────────────────────────────────────

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def scrape_sales(page):
    """P&C Sales by Producer — returns (headers, rows)."""
    url = "https://app.agencyzoom.com/sales-report/index#"
    headers = ["Producer", "Premium", "Policies", "Items", "Sales", "Revenue"]

    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.wait_for_selector("#detail-container table tbody tr.line", timeout=30_000)
    except PlaywrightTimeoutError:
        return headers, []

    rows_data = []
    for row in page.query_selector_all("#detail-container table tbody tr.line.parent"):
        def cv(sel):
            el = row.query_selector(sel)
            return (el.get_attribute("data-val") or "").strip() if el else ""

        def fmt(val):
            try:
                return f"${int(val):,}"
            except (ValueError, TypeError):
                return val

        producer_td = row.query_selector("td.lsp")
        rows_data.append([
            (producer_td.get_attribute("data-val") or "").strip(),
            fmt(cv("td.premium")),
            cv("td.policies"),
            cv("td.items"),
            cv("td.sales"),
            fmt(cv("td.commission")),
        ])
    return headers, rows_data


def scrape_monthly(page):
    """Monthly Sales Summary — returns (headers, rows)."""
    url = "https://app.agencyzoom.com/report/index?report=monthly"
    headers = ["Month",
               "PC_Sales", "PC_Policies", "PC_Items", "PC_Premium",
               "LH_Appts", "LH_Policies", "LH_Premium"]

    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.wait_for_selector("#detail-container table tbody tr.parent", timeout=30_000)
    except PlaywrightTimeoutError:
        return headers, []

    rows_data = []
    for row in page.query_selector_all("#detail-container table tbody tr.parent"):
        tds = row.query_selector_all("td")
        def t(i):
            return tds[i].inner_text().strip() if i < len(tds) else ""
        rows_data.append([t(0), t(1), t(2), t(3), t(4), t(5), t(6), t(7)])
    return headers, rows_data


def scrape_annual(page):
    """Annual Sales Summary — returns (headers, rows, total_row_index)."""
    url = "https://app.agencyzoom.com/report/index?report=annual"
    headers = (
        ["Staff"]
        + [f"{m}_{t}" for m in MONTHS for t in ("Items", "Premium")]
        + ["Year_Items", "Year_Premium"]
    )

    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.wait_for_selector("#detail-container table tbody tr", timeout=30_000)
    except PlaywrightTimeoutError:
        return headers, [], -1

    rows_data = []
    total_row_index = -1
    for i, row in enumerate(page.query_selector_all("#detail-container table tbody tr")):
        first = row.query_selector("th, td")
        if not first:
            continue
        staff = first.inner_text().strip()
        is_total = first.evaluate("el => el.tagName") == "TH"
        if is_total:
            total_row_index = i

        tds = row.query_selector_all("td")
        data_cells = tds if is_total else tds[1:]

        def cell(idx):
            return data_cells[idx].inner_text().strip() if idx < len(data_cells) else ""

        record = [staff]
        col = 0
        for _ in MONTHS:
            record.append(cell(col));   col += 1
            record.append(cell(col));   col += 1
        record.append(cell(col));   col += 1
        record.append(cell(col))
        rows_data.append(record)
    return headers, rows_data, total_row_index


# ── pipeline scrapers ───────────────────────────────────────────────────────

@dataclass(frozen=True)
class PipelineConfig:
    """Per-section knobs that differ between lead and service pipelines."""
    list_url: str           # page hosting the pipeline-selector dropdown
    pipeline_url: str       # page rendering a single pipeline's columns
    base_path: str          # path fragment that pipeline links must contain
    count_selector: str     # CSS for the count h4 inside a column header
    count_label: str        # human-readable header for the count column


LEAD_PIPELINES = PipelineConfig(
    list_url='https://app.agencyzoom.com/referral/pipeline',
    pipeline_url='https://app.agencyzoom.com/referral/pipeline',
    base_path='/referral/pipeline',
    count_selector='.dd-header-counter h4[title="Number of Leads"]',
    count_label='Lead Count',
)

SERVICE_PIPELINES = PipelineConfig(
    list_url='https://app.agencyzoom.com/pipeline/service-pipeline',
    pipeline_url='https://app.agencyzoom.com/pipeline/service',
    base_path='/pipeline/service',
    count_selector='.dd-header-counter h4[id^="count"]',
    count_label='Ticket Count',
)


def _discover_pipelines(page, list_url, base_path):
    """Return [(name, workflow_id), ...] for pipelines linked from list_url."""
    page.goto(list_url, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.wait_for_selector(".dropdown.pipelineDropdown", timeout=30_000)
    except PlaywrightTimeoutError:
        return []

    selector = (
        ".dropdown.pipelineDropdown .dropdown-menu "
        "a.dropdown-item:not(.dropdown-item__main):not(.dropdown-item__action)"
    )
    pipelines = []
    seen = set()
    for link in page.query_selector_all(selector):
        href = link.get_attribute("href") or ""
        name = (link.inner_text() or "").strip()
        if base_path not in href or "workflowId=" not in href:
            continue
        wf = parse_qs(urlparse(href).query).get("workflowId", [""])[0]
        if not wf or wf in seen:
            continue
        seen.add(wf)
        pipelines.append((name, wf))
    return pipelines


def _scrape_pipeline_stages(page, workflow_id, pipeline_url, count_selector):
    """Return [(stage_name, count), ...] for one pipeline."""
    url = f"{pipeline_url}?workflowId={workflow_id}"
    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.wait_for_selector(".dd-column", timeout=30_000)
    except PlaywrightTimeoutError:
        return []

    # Counts populate via async update after the columns mount; settle briefly.
    page.wait_for_timeout(1500)

    stages = []
    for column in page.query_selector_all(".dd-column"):
        name_el = column.query_selector(".dd-header h2")
        count_el = column.query_selector(count_selector)
        if not name_el:
            continue
        stage_name = name_el.inner_text().strip()
        count_text = (count_el.inner_text().strip() if count_el else "").replace(",", "")
        try:
            count = int(count_text) if count_text else 0
        except ValueError:
            count = count_text
        stages.append((stage_name, count))
    return stages


def scrape_all_pipelines(page, config, status_cb=print):
    """Discover every pipeline for `config` and scrape each one's stages.

    Returns [(pipeline_name, [(stage_name, count), ...]), ...].
    """
    pipelines = _discover_pipelines(page, config.list_url, config.base_path)
    status_cb(f"  Found {len(pipelines)} pipeline(s).")
    results = []
    for name, wf in pipelines:
        status_cb(f"  '{name}' (workflowId={wf})...")
        stages = _scrape_pipeline_stages(
            page, wf, config.pipeline_url, config.count_selector)
        status_cb(f"    {len(stages)} stage(s) found.")
        results.append((name, stages))
    return results


# ── Excel: styling ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL  = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT  = Font(bold=True)
TITLE_FONT  = Font(bold=True, size=12, color="1F4E79")


def _style_header_row(ws, row_num, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _style_total_row(ws, row_num, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT


def _style_title_cell(ws, row_num):
    ws.cell(row=row_num, column=1).font = TITLE_FONT


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


# ── Excel: writing ──────────────────────────────────────────────────────────

BLOCK_GAP = 4   # blank rows between stacked blocks


def write_sheet(ws, headers, rows, total_row_index=-1):
    """Write a single header + data block to a sheet."""
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for i, row in enumerate(rows):
        ws.append(row)
        if i == total_row_index:
            _style_total_row(ws, i + 2, len(headers))
    _autofit(ws)


def write_stacked_sheet(ws, blocks):
    """Render `blocks` vertically with BLOCK_GAP blank rows between each.

    Each block is a dict: {title, headers, rows, total_row_index (optional)}.
    """
    next_row = 1
    for b in blocks:
        title    = b["title"]
        headers  = b["headers"]
        rows     = b["rows"]
        total_i  = b.get("total_row_index", -1)

        ws.cell(row=next_row, column=1, value=title)
        _style_title_cell(ws, next_row)
        header_row = next_row + 1

        for c, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=c, value=h)
        _style_header_row(ws, header_row, len(headers))

        for i, row in enumerate(rows):
            target_row = header_row + 1 + i
            for c, val in enumerate(row, start=1):
                ws.cell(row=target_row, column=c, value=val)
            if i == total_i:
                _style_total_row(ws, target_row, len(headers))

        last_row = header_row + len(rows)
        next_row = last_row + 1 + BLOCK_GAP

    _autofit(ws)


# ── Excel: sheet-name helpers ───────────────────────────────────────────────

_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def sanitize_sheet_name(name):
    """Strip Excel-reserved chars and clamp to 31 characters."""
    cleaned = _INVALID_SHEET_CHARS.sub("", name).strip()
    return cleaned[:31] or "Sheet"


def make_unique_sheet_name(name, used):
    """Return a sanitized, deduped sheet name. Mutates `used`."""
    base = sanitize_sheet_name(name)
    candidate = base
    n = 2
    while candidate in used:
        candidate = f"{base[:28]} {n}"
        n += 1
    used.add(candidate)
    return candidate
