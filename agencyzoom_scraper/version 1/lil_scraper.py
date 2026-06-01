"""
Lil Scraper — GUI front-end that runs every AgencyZoom scrape in one
session and writes a single Excel file with three sheets:

  Reports           — Sales, Monthly, and Annual reports stacked
  Lead Pipelines    — every lead pipeline stacked
  Service Pipelines — every service pipeline stacked

Pipelines pull every lead/ticket regardless of date. (The date-range
"Filtered Scrape" mode has been archived — see
beta old scrape/custom_scrape_feature.py.)

Output: agencyzoom_all_data.xlsx
"""

import threading
import tkinter as tk
from tkinter import messagebox
from pathlib import Path

from openpyxl import Workbook

from scraper_core import (
    scraper_session,
    scrape_sales, scrape_monthly, scrape_annual,
    scrape_all_pipelines, LEAD_PIPELINES, SERVICE_PIPELINES,
    write_stacked_sheet,
)

SCRAPER_PROFILE = Path(__file__).parent / "scraper_profile"
OUTPUT_FILE     = Path(__file__).parent / "agencyzoom_all_data.xlsx"


def run_scrape(status_cb, login_prompt):
    """Scrape everything in one browser session and write the combined workbook."""
    with scraper_session(SCRAPER_PROFILE,
                         login_prompt=login_prompt,
                         status_cb=status_cb) as (page, _):
        status_cb("Scraping Sales report...")
        sales_h,   sales_rows                  = scrape_sales(page)
        status_cb("Scraping Monthly report...")
        monthly_h, monthly_rows                = scrape_monthly(page)
        status_cb("Scraping Annual report...")
        annual_h,  annual_rows, annual_total_i = scrape_annual(page)
        status_cb("Scraping lead pipelines...")
        lead_results = scrape_all_pipelines(page, LEAD_PIPELINES,
                                            status_cb=status_cb)
        status_cb("Scraping service pipelines...")
        service_results = scrape_all_pipelines(page, SERVICE_PIPELINES,
                                               status_cb=status_cb)

    status_cb("Writing workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    write_stacked_sheet(wb.create_sheet("Reports"), [
        {"title": "Sales Report (P&C Sales by Producer)",
         "headers": sales_h, "rows": sales_rows},
        {"title": "Monthly Sales Summary",
         "headers": monthly_h, "rows": monthly_rows},
        {"title": "Annual Sales Summary",
         "headers": annual_h, "rows": annual_rows,
         "total_row_index": annual_total_i},
    ])

    write_stacked_sheet(wb.create_sheet("Lead Pipelines"),
                        _pipeline_blocks(lead_results, LEAD_PIPELINES.count_label))
    write_stacked_sheet(wb.create_sheet("Service Pipelines"),
                        _pipeline_blocks(service_results, SERVICE_PIPELINES.count_label))

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


def _pipeline_blocks(results, count_label):
    """Convert [(name, stages), ...] into block dicts for write_stacked_sheet."""
    if not results:
        return [{"title": "(none found)",
                 "headers": ["Stage", count_label], "rows": []}]
    return [
        {"title": name,
         "headers": ["Stage", count_label],
         "rows": [list(s) for s in stages]}
        for name, stages in results
    ]


# ── GUI ──────────────────────────────────────────────────────────────────────

def main():
    root = tk.Tk()
    root.title("Lil Scraper")
    root.geometry("380x250")
    root.resizable(False, False)

    frame = tk.Frame(root, padx=20, pady=20)
    frame.pack(expand=True, fill="both")

    tk.Label(frame, text="Lil Scraper",
             font=("Segoe UI", 16, "bold")).pack(pady=(0, 10))

    status_var = tk.StringVar(value="Ready.")
    tk.Label(frame, textvariable=status_var, font=("Segoe UI", 9),
             wraplength=340, justify="center").pack(pady=(0, 12))

    def status_cb(msg):
        root.after(0, lambda: status_var.set(msg))

    def login_prompt(browser_name):
        done = threading.Event()

        def show():
            messagebox.showinfo(
                "Login required",
                f"Log in to AgencyZoom in the {browser_name} window,\n"
                f"then click OK to continue."
            )
            done.set()

        root.after(0, show)
        done.wait()

    def start_worker():
        scrape_btn.config(state="disabled")

        def worker():
            try:
                out = run_scrape(status_cb, login_prompt)
                status_cb(f"Done — saved to {out.name}")
                root.after(0, lambda: messagebox.showinfo(
                    "Lil Scraper", f"Done!\nSaved to:\n{out}"))
            except Exception as e:
                err = str(e)
                status_cb(f"Error: {err}")
                root.after(0, lambda: messagebox.showerror("Lil Scraper", err))
            finally:
                root.after(0, lambda: scrape_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()

    btn_frame = tk.Frame(frame)
    btn_frame.pack()
    scrape_btn = tk.Button(btn_frame, text="Scrape", width=18,
                           font=("Segoe UI", 10), command=start_worker)
    scrape_btn.pack(pady=2)

    root.mainloop()


if __name__ == "__main__":
    main()
