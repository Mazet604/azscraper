"""
Scrapes all three AgencyZoom reports in one run and writes them
to a single Excel file with one sheet per report.

Sheets:
  Sales Report   — P&C Sales by Producer
  Monthly Report — Monthly Sales Summary
  Annual Report  — Annual Sales Summary

Brave must be running with --remote-debugging-port=9222.
No need to close Brave — the script opens a new tab and closes only that tab.
Output: agencyzoom_reports.xlsx
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

DEBUG_PORT   = 9222
OUTPUT_FILE  = Path(__file__).parent / "agencyzoom_reports.xlsx"

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun",
          "Jul","Aug","Sep","Oct","Nov","Dec"]

# ── styling helpers ───────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL  = PatternFill("solid", fgColor="D9E1F2")   # light blue
TOTAL_FONT  = Font(bold=True)


def style_header(ws, row_num, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def style_total(ws, row_num, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


# ── scrapers ──────────────────────────────────────────────────────────────────

def scrape_sales(page):
    """P&C Sales by Producer — returns (headers, rows)."""
    url = "https://app.agencyzoom.com/sales-report/index#"
    headers = ["Producer", "Premium", "Policies", "Items", "Sales", "Revenue"]

    print(f"  Navigating to Sales Report...")
    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.wait_for_selector("#detail-container table tbody tr.line", timeout=30_000)
    except PlaywrightTimeoutError:
        print("  ERROR: Sales report table did not load.")
        return headers, []

    rows_data = []
    for row in page.query_selector_all("#detail-container table tbody tr.line.parent"):
        def cv(sel):
            el = row.query_selector(sel)
            return (el.get_attribute("data-val") or "").strip() if el else ""

        def fmt(val):
            try:
                return f"${int(val):,}"
            except (ValueError, TypeError):
                return val

        producer_td = row.query_selector("td.lsp")
        rows_data.append([
            (producer_td.get_attribute("data-val") or "").strip(),
            fmt(cv("td.premium")),
            cv("td.policies"),
            cv("td.items"),
            cv("td.sales"),
            fmt(cv("td.commission")),
        ])

    print(f"  Sales Report: {len(rows_data)} producers found.")
    return headers, rows_data


def scrape_monthly(page):
    """Monthly Sales Summary — returns (headers, rows)."""
    url = "https://app.agencyzoom.com/report/index?report=monthly"
    headers = ["Month",
               "PC_Sales", "PC_Policies", "PC_Items", "PC_Premium",
               "LH_Appts", "LH_Policies", "LH_Premium"]

    print(f"  Navigating to Monthly Report...")
    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.wait_for_selector(
            "#detail-container table tbody tr.parent", timeout=30_000)
    except PlaywrightTimeoutError:
        print("  ERROR: Monthly report table did not load.")
        return headers, []

    rows_data = []
    for row in page.query_selector_all("#detail-container table tbody tr.parent"):
        tds = row.query_selector_all("td")
        def t(i):
            return tds[i].inner_text().strip() if i < len(tds) else ""
        rows_data.append([t(0), t(1), t(2), t(3), t(4), t(5), t(6), t(7)])

    print(f"  Monthly Report: {len(rows_data)} months found.")
    return headers, rows_data


def scrape_annual(page):
    """Annual Sales Summary — returns (headers, rows, total_row_index)."""
    url = "https://app.agencyzoom.com/report/index?report=annual"
    headers = (
        ["Staff"]
        + [f"{m}_{t}" for m in MONTHS for t in ("Items", "Premium")]
        + ["Year_Items", "Year_Premium"]
    )

    print(f"  Navigating to Annual Report...")
    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.wait_for_selector(
            "#detail-container table tbody tr", timeout=30_000)
    except PlaywrightTimeoutError:
        print("  ERROR: Annual report table did not load.")
        return headers, [], -1

    rows_data = []
    total_row_index = -1
    for i, row in enumerate(
        page.query_selector_all("#detail-container table tbody tr")
    ):
        first = row.query_selector("th, td")
        if not first:
            continue
        staff = first.inner_text().strip()
        is_total = first.evaluate("el => el.tagName") == "TH"
        if is_total:
            total_row_index = i

        tds = row.query_selector_all("td")
        data_cells = tds if is_total else tds[1:]

        def cell(idx):
            return data_cells[idx].inner_text().strip() if idx < len(data_cells) else ""

        record = [staff]
        col = 0
        for _ in MONTHS:
            record.append(cell(col));   col += 1
            record.append(cell(col));   col += 1
        record.append(cell(col));   col += 1
        record.append(cell(col))
        rows_data.append(record)

    print(f"  Annual Report: {len(rows_data)} staff rows found.")
    return headers, rows_data, total_row_index


# ── write workbook ────────────────────────────────────────────────────────────

def write_sheet(ws, headers, rows, total_row_index=-1):
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for i, row in enumerate(rows):
        ws.append(row)
        if i == total_row_index:
            style_total(ws, i + 2, len(headers))
    autofit(ws)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    with sync_playwright() as p:
        print(f"Connecting to Brave via CDP on port {DEBUG_PORT}...")
        try:
            browser = p.chromium.connect_over_cdp(
                f"http://localhost:{DEBUG_PORT}",
                timeout=15_000,
            )
        except Exception as e:
            print(f"ERROR: Could not connect to Brave — {e}")
            print(
                "\nMake sure Brave is running with remote debugging enabled.\n"
                "Right-click your Brave shortcut → Properties → add to Target:\n"
                '  --remote-debugging-port=9222\n'
                "Then restart Brave."
            )
            sys.exit(1)

        context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = context.new_page()

        print("\nScraping all reports (Brave will stay open while navigating)...")

        sales_headers,   sales_rows            = scrape_sales(page)
        monthly_headers, monthly_rows          = scrape_monthly(page)
        annual_headers,  annual_rows, total_idx = scrape_annual(page)

        page.close()

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    ws_sales = wb.create_sheet("Sales Report")
    write_sheet(ws_sales, sales_headers, sales_rows)

    ws_monthly = wb.create_sheet("Monthly Report")
    write_sheet(ws_monthly, monthly_headers, monthly_rows)

    ws_annual = wb.create_sheet("Annual Report")
    write_sheet(ws_annual, annual_headers, annual_rows, total_idx)

    wb.save(OUTPUT_FILE)
    print(f"\nAll done! Saved to: {OUTPUT_FILE}")
    print(f"  Sheet 'Sales Report'   — {len(sales_rows)} rows")
    print(f"  Sheet 'Monthly Report' — {len(monthly_rows)} rows")
    print(f"  Sheet 'Annual Report'  — {len(annual_rows)} rows")


if __name__ == "__main__":
    main()
